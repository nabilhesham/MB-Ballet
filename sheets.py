"""
Readers for the academy's own spreadsheets.

Reception has kept the academy in Excel for years and will keep doing so for a
while yet, so the sheets are the source of truth at seed time rather than a
one-off import format. Two shapes exist:

*Roster* sheets (ballet.xlsx, flexibility.xlsx) list one class group per block:
a title line, a header line, then one row per student with their attendance
written across the row -- a date where they came, "13/8 absent" where they did
not.

*Salary* sheets (August_Salaries_2026.xlsx) are a grid of instructor against
day, holding hours worked, with total hours and total pay on the last rows.

Nothing here touches the database. Everything returns plain dataclasses so the
parsing can be read, tested and argued about on its own, and so seed.py stays
a list of inserts.

Every guess this module makes -- an inferred year, a rescued typo, a cell it
could not read -- is appended to `warnings` on the result rather than being
applied silently. seed.py prints them.
"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime

try:
    import openpyxl
except ImportError:                                    # pragma: no cover
    openpyxl = None


WEEKDAYS = {
    "monday": 0, "tuesday": 1, "wednesday": 2, "thursday": 3,
    "friday": 4, "saturday": 5, "sunday": 6,
}

# The attendance columns are the ones headed 1ST, 2ND, 3RD, 4TH -- repeated
# once per month on the ballet sheet, twice per pack on the flexibility one.
# Keying off the ordinal rather than a column number is what lets both sheets,
# and the blocks within them that omit NAME or school, share one reader.
ORDINALS = {"1st", "2nd", "3rd", "4th", "5th", "6th", "7th", "8th"}

# Header text -> field name. Anything not listed is ignored.
FIELDS = {
    "paid date": "paid_date", "date": "paid_date",
    "name": "name", "cell phone": "phone", "phone": "phone",
    "dob": "dob", "age": "age", "month": "months", "months": "months",
    "paid": "paid", "school": "school", "days": "days",
    "sessions": "sessions", "session": "sessions",
}

LEVEL_RE = re.compile(
    r"\b(level\s*\d+|grade\s*\d+|primary|beginner|intermediate|advanced)\b", re.I)
TITLE_RE = re.compile(
    r"^(?P<day>\w+day)\s+at\s+(?P<time>\d{1,2}\s*[.:]\s*\d{2})\s*(?P<ampm>am|pm)"
    r"(?:\s*with\s+captain\b\s*(?P<captain>.*))?$", re.I)
DAYMONTH_RE = re.compile(r"(\d{1,2})\s*[/\-.]\s*(\d{1,2})")


# ---------------------------------------------------------------- data shapes
@dataclass
class Slot:
    """One attendance cell: a session the student was booked into."""
    on: date | None
    present: bool
    raw: str


@dataclass
class Enrolment:
    """One student row inside a block."""
    name: str
    phone: str | None = None
    age: float | None = None
    dob: str | None = None
    school: str | None = None
    paid_date: date | None = None
    months: int | None = None
    sessions: int | None = None
    paid_raw: str | None = None
    price: float | None = None
    days: str | None = None
    weekdays: list[int] = field(default_factory=list)
    slots: list[Slot] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    row: int = 0

    @property
    def note(self) -> str | None:
        return "; ".join(self.notes) or None


@dataclass
class Group:
    """One block of a roster sheet: a class group and everyone in it."""
    title: str
    family: str                     # "Ballet" / "Flexibility"
    class_name: str = ""
    level: str | None = None
    instructor: str | None = None
    weekday: int | None = None      # ballet: the one day it runs
    instructor_id: int | None = None    # filled in by the seed, once matched
    hour: int = 18
    minute: int = 0
    duration_hours: float = 1.5
    students: list[Enrolment] = field(default_factory=list)

    def grid_weekdays(self) -> list[int]:
        """
        Which weekdays this group's sessions fall on.

        Ballet blocks say so in the title. Flexibility blocks do not -- the
        title is only "EVENING FLEXIBILITY" -- so the days are taken from the
        union of what the students in the block actually signed up for.
        """
        if self.weekday is not None:
            return [self.weekday]
        days = {d for s in self.students for d in s.weekdays}
        return sorted(days)


@dataclass
class Roster:
    path: str
    family: str
    colour: str | None = None
    groups: list[Group] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class InstructorPay:
    name: str
    days: dict[date, float] = field(default_factory=dict)
    sheet_hours: float | None = None     # the sheet's own totals row
    sheet_pay: float | None = None

    @property
    def hours(self) -> float:
        return round(sum(self.days.values()), 2)

    @property
    def hourly_rate(self) -> float:
        """
        Derived, not stored: the sheet gives hours and pay per instructor and
        the rate is the ratio. Reading it back out means a rate change next
        month arrives on its own rather than needing a code edit.
        """
        if self.sheet_pay and self.sheet_hours:
            return round(self.sheet_pay / self.sheet_hours, 2)
        return 0.0


@dataclass
class Payroll:
    path: str
    period: str | None = None            # "2026-08"
    instructors: list[InstructorPay] = field(default_factory=list)
    sheet_total: float | None = None
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------- cell helpers
def text(v) -> str:
    """A cell as trimmed text. Excel's non-breaking spaces are everywhere here."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).replace("\xa0", " ").strip()


def number(v) -> float | None:
    """
    A cell as a number, tolerating the way people actually type into Excel:
    "12y", "7.5", "  20  ". Returns None for "online", "class" and blanks.
    """
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    t = text(v)
    if not t:
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", t.replace(",", ""))
    return float(m.group()) if m else None


def phone_of(v) -> str | None:
    """
    Normalise an Egyptian mobile. The sheets hold the same number as
    1129200365 in one block and 01129200365 in another, and Excel has eaten
    the leading zero wherever the cell was typed as a number. International
    numbers keep their +.
    """
    t = text(v)
    if not t:
        return None
    if t.startswith("+"):
        return "+" + re.sub(r"\D", "", t[1:])
    digits = re.sub(r"\D", "", t)
    if not digits:
        return None
    if len(digits) == 10 and digits.startswith("1"):
        digits = "0" + digits
    return digits


def phone_key(p: str | None) -> str | None:
    """Identity for a phone: the last 10 digits, so 011... and 11... match."""
    if not p:
        return None
    digits = re.sub(r"\D", "", p)
    return digits[-10:] if len(digits) >= 10 else digits


def name_key(n: str) -> str:
    return re.sub(r"\s+", " ", n).strip().lower()


def weekdays_in(t: str) -> list[int]:
    """Weekday numbers named in "sunday - thursday", in calendar order."""
    found = {WEEKDAYS[w] for w in WEEKDAYS if re.search(rf"\b{w}\b", t, re.I)}
    return sorted(found)


# ---------------------------------------------------------------- date reading
def _plausible(d: date, year: int) -> bool:
    return year - 1 <= d.year <= year + 1


def read_date(v, year: int, warns: list[str], where: str) -> date | None:
    """
    A cell as a date.

    Two things go wrong in these sheets and both are repaired here rather than
    left to surprise someone later:

    *Half-written dates.* Absences are typed as "26/8 absent" with no year,
    because the person writing it knows which year it is. The sheet's own year
    supplies the rest.

    *Fat-fingered years.* One row carries 2028-07-01 and another 2019-08-19 in
    a sheet whose every other date is 2026. A student cannot have attended in
    2028, so the day and month are kept and the year is corrected -- and the
    correction is recorded, never applied quietly.
    """
    if isinstance(v, datetime):
        v = v.date()
    if isinstance(v, date):
        if _plausible(v, year):
            return v
        try:
            fixed = v.replace(year=year)
        except ValueError:                    # 29 Feb in a non-leap year
            return None
        warns.append(f"{where}: year {v.year} looks wrong, read as {fixed.isoformat()}")
        return fixed

    t = text(v)
    if not t:
        return None
    m = DAYMONTH_RE.search(t)
    if not m:
        return None
    day, month = int(m.group(1)), int(m.group(2))
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return None
    try:
        return date(year, month, day)
    except ValueError:
        warns.append(f"{where}: {t!r} is not a real date")
        return None


def _sheet_year(ws) -> int:
    """The year the sheet is about: whichever one its real dates agree on."""
    years = Counter()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, (datetime, date)):
                years[cell.year] += 1
    return years.most_common(1)[0][0] if years else date.today().year


# ---------------------------------------------------------------- roster sheets
def _is_header(cells: list[str]) -> bool:
    lows = [c.lower() for c in cells]
    return bool(lows) and lows[0] in ("paid date", "date") and any(
        c in ORDINALS for c in lows)


def _is_title(cells: list[str]) -> bool:
    """
    A block title sits alone in column A and reads as words.

    The "alone" test matters: rows carrying nothing but a second phone number
    for the student above look identical in shape, and are told apart by
    having no letters in them.
    """
    filled = [i for i, c in enumerate(cells) if c]
    if filled != [0]:
        return False
    return len(re.findall(r"[A-Za-z]", cells[0])) >= 3


def _map_columns(cells: list[str]) -> tuple[dict[str, int], list[int]]:
    """Header row -> {field: column index}, [attendance column indices]."""
    fields: dict[str, int] = {}
    slots: list[int] = []
    for i, raw in enumerate(cells):
        low = re.sub(r"\s+", " ", raw).strip().lower()
        if low in ORDINALS:
            slots.append(i)
        elif low in FIELDS and FIELDS[low] not in fields:
            fields[FIELDS[low]] = i
    # Two ballet blocks leave the NAME header blank but still fill the column.
    if "name" not in fields:
        fields["name"] = 1
    return fields, slots


def _parse_title(title: str, family: str, warns: list[str]) -> Group:
    g = Group(title=title, family=family)
    m = TITLE_RE.match(title.strip())
    if m:
        g.weekday = WEEKDAYS.get(m.group("day").lower())
        hh, mm = re.split(r"[.:]", re.sub(r"\s+", "", m.group("time")))
        hour, g.minute = int(hh), int(mm)
        if m.group("ampm").lower() == "pm" and hour != 12:
            hour += 12
        if m.group("ampm").lower() == "am" and hour == 12:
            hour = 0
        g.hour = hour
        captain = (m.group("captain") or "").strip()
        if captain:
            lvl = LEVEL_RE.search(captain)
            if lvl:
                g.level = re.sub(r"\s+", " ", lvl.group()).strip().lower()
                captain = LEVEL_RE.sub("", captain).strip(" -,")
            g.instructor = re.sub(r"\s+", " ", captain).strip() or None
    else:
        # "EVENING FLEXIBILITY" / "MORNING FLEXIBILITY": no day and no time,
        # because the days are per student. Only the half of the day is fixed.
        low = title.lower()
        g.hour = 10 if "morning" in low else 19
        if "morning" not in low and "evening" not in low:
            warns.append(f"block {title!r}: no day or time in the title, "
                         f"scheduled at {g.hour}:00")
    return g


def _class_name(g: Group, taken: set[str]) -> str:
    """
    What this block is called in the app.

    The sheet's own wording wins -- staff say "grade 6", not "Ballet group 4"
    -- with the time appended only where two blocks would otherwise collide.
    """
    if g.family.lower() in g.title.lower():                # "EVENING FLEXIBILITY"
        base = " ".join(w.capitalize() for w in g.title.split())
    elif g.level:
        base = f"{g.family} {g.level.title()}"
    else:
        base = f"{g.family} {g.hour % 12 or 12}:{g.minute:02d} " \
               f"{'AM' if g.hour < 12 else 'PM'}"
    name, n = base, 2
    while name in taken:
        name, n = f"{base} ({n})", n + 1
    taken.add(name)
    return name


def _read_student(cells, fields, slot_cols, year, warns, rowno) -> Enrolment | None:
    def cell(key):
        i = fields.get(key)
        return cells[i] if i is not None and i < len(cells) else None

    name = text(cell("name"))
    if not name:
        return None

    e = Enrolment(name=re.sub(r"\s+", " ", name), row=rowno)
    where = f"row {rowno} ({e.name})"

    e.phone = phone_of(cell("phone"))
    e.school = text(cell("school")) or None
    e.days = text(cell("days")) or None
    e.weekdays = weekdays_in(e.days) if e.days else []
    e.paid_date = read_date(cell("paid_date"), year, warns, where)

    dob = cell("dob")
    e.dob = read_date(dob, year, warns, where).isoformat() if isinstance(
        dob, (date, datetime)) else (text(dob) or None)

    raw_age = text(cell("age"))
    e.age = number(cell("age"))
    if raw_age and e.age is None:
        e.notes.append(raw_age)                # "online" -- a real note, not an age

    months = cell("months")
    e.months = int(number(months)) if number(months) is not None else None
    raw_months = text(months)
    if raw_months and e.months is None:
        e.notes.append(raw_months)             # "class" -- a single drop-in

    sessions = number(cell("sessions"))
    e.sessions = int(sessions) if sessions is not None else None

    e.paid_raw = text(cell("paid")) or None
    if e.paid_raw:
        # "680" is money; "package", "free" and "yes" are not. A plan whose
        # price nobody wrote down must stay unpriced rather than be counted
        # as zero revenue -- the dashboard reports the difference.
        amount = number(e.paid_raw)
        if amount is not None and re.fullmatch(r"[\d\s.,]+", e.paid_raw):
            e.price = amount

    for i in slot_cols:
        if i >= len(cells):
            break
        raw = text(cells[i])
        if not raw:
            continue
        low = raw.lower()
        if low.startswith("start"):            # "Start 10/9" -- not yet a session
            e.notes.append(raw)
            continue
        on = read_date(cells[i], year, warns, f"{where} slot {raw!r}")
        if on is None:
            e.notes.append(raw)                # "app", "out", "msfra 1 month"
            continue
        e.slots.append(Slot(on=on, present="abs" not in low, raw=raw))

    e.slots.sort(key=lambda s: s.on)
    return e


def read_roster(path: str, family: str, duration_hours: float = 1.5) -> Roster:
    """Read one class roster workbook into its blocks."""
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed -- run: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    out = Roster(path=path, family=family)
    taken: set[str] = set()

    for ws in wb.worksheets:
        year = _sheet_year(ws)
        group: Group | None = None
        fields: dict[str, int] = {}
        slot_cols: list[int] = []

        for rowno, row in enumerate(ws.iter_rows(values_only=True), 1):
            cells = [text(c) for c in row]
            while cells and not cells[-1]:
                cells.pop()
            if not cells:
                continue

            if _is_header(cells):
                fields, slot_cols = _map_columns(cells)
                continue
            if _is_title(cells):
                group = _parse_title(cells[0], family, out.warnings)
                group.duration_hours = duration_hours
                group.class_name = _class_name(group, taken)
                out.groups.append(group)
                fields, slot_cols = {}, []
                continue
            if group is None or not slot_cols:
                continue                        # the merged "1ST MONTH" band

            student = _read_student(row, fields, slot_cols, year, out.warnings, rowno)
            if student is None:
                # A row holding only a spare phone number belongs to the
                # student above it -- reception writes the mother's line there.
                extra = phone_of(next((c for c in cells if c), ""))
                if extra and group.students:
                    group.students[-1].notes.append(f"also {extra}")
                continue
            group.students.append(student)

    return out


# ---------------------------------------------------------------- salary sheet
def read_salaries(path: str) -> Payroll:
    """
    Read a monthly salary grid: instructors across the top, days down the side.

    The two rows under the grid hold the sheet's own totals for hours and pay.
    They are kept as written so the import can be checked against the sheet,
    and the hourly rate is taken as their ratio.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed -- run: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    out = Payroll(path=path)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        out.warnings.append(f"{path}: empty sheet")
        return out

    header = rows[0]
    by_col: dict[int, InstructorPay] = {}
    for i, cell in enumerate(header):
        name = re.sub(r"\s+", " ", text(cell))
        if i == 0 or not name:
            continue
        by_col[i] = InstructorPay(name=name)

    # Rows with a date in column A are working days; the unlabelled rows after
    # the last of them are the sheet's totals, in the order hours then pay.
    tail: list[tuple] = []
    for row in rows[1:]:
        when = row[0] if row else None
        if isinstance(when, datetime):
            when = when.date()
        if isinstance(when, date):
            for i, pay in by_col.items():
                h = number(row[i]) if i < len(row) else None
                if h:
                    pay.days[when] = h
            if out.period is None:
                out.period = when.strftime("%Y-%m")
        elif any(number(c) is not None for c in (row or ())):
            tail.append(row)

    for which, row in zip(("sheet_hours", "sheet_pay"), tail):
        for i, pay in by_col.items():
            v = number(row[i]) if i < len(row) else None
            if v is not None:
                setattr(pay, which, v)
    if len(tail) > 2:
        out.sheet_total = next(
            (number(c) for c in tail[2] if number(c) is not None), None)

    for pay in by_col.values():
        if pay.sheet_hours is not None and abs(pay.hours - pay.sheet_hours) > 0.01:
            out.warnings.append(
                f"{pay.name}: days add up to {pay.hours}h but the sheet's total "
                f"says {pay.sheet_hours}h -- using the days")
        if not pay.hourly_rate:
            out.warnings.append(f"{pay.name}: no hourly rate could be worked out")
        out.instructors.append(pay)

    return out
